<<<<<<< HEAD
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook

#엑셀 데이터 불러오기
load_wb = load_workbook("C:/Users/GL/Desktop/application_info.xlsx", data_only=True)
#시트 이름으로 불러오기
load_ws = load_wb['Sheet1']

#열린 크롬창 제어하기
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_driver = "C:/Users/GL/Desktop/chromedriver.exe" # Your Chrome Driver path
driver = webdriver.Chrome(chrome_driver, options=chrome_options)

#사이트접속
driver.get(str(load_ws['B1'].value))

#영문이름
elem = driver.find_element_by_id("englishName")
elem.send_keys(load_ws['B6'].value)

#한문이름
elem = driver.find_element_by_id("chineseName")
elem.send_keys(load_ws['B7'].value)

#성별 #1
# elem = driver.find_elements_by_class_name("label")
# elem[0].click()

#성별 #2
elem = driver.find_element_by_xpath('//*[@id="wrapResume"]/section[1]/div[1]/div[3]/div/label[1]/span')
elem.click()

#생년월일
elem = driver.find_element_by_id("birthday")
elem.send_keys(load_ws['B8'].value)

#빈 곳 클릭
elem = driver.find_element_by_class_name("h1")
elem.click()

#병역사항 #1
# elem = driver.find_elements_by_class_name("label")
# elem[5].click()

#병역사항 #2
elem = driver.find_element_by_link_text("군필")
elem.click()

# elem = driver.find_element_by_id("militaryRole")
# elem.send_keys(load_ws['B9'].value)

# elem = driver.find_element_by_name("military.militaryStartDate")
# elem.send_keys(load_ws['B10'].value)

# elem = driver.find_element_by_name("military.militaryEndDate")
=======
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook

#엑셀 데이터 불러오기
load_wb = load_workbook("C:/Users/GIGM/Desktop/application_info.xlsx", data_only=True)
#시트 이름으로 불러오기
load_ws = load_wb['Sheet1']

#열린 크롬창 제어하기
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_driver = "C:/Users/GIGM/Desktop/chromedriver.exe" # Your Chrome Driver path
driver = webdriver.Chrome(chrome_driver, options=chrome_options)

#사이트접속
driver.get(str(load_ws['B1'].value))

#영문이름
elem = driver.find_element_by_id("englishName")
elem.send_keys(load_ws['B6'].value)

#한문이름
elem = driver.find_element_by_id("chineseName")
elem.send_keys(load_ws['B7'].value)

#성별 #1
# elem = driver.find_elements_by_class_name("label")
# elem[0].click()

#성별 #2
elem = driver.find_element_by_xpath('//*[@id="wrapResume"]/section[1]/div[1]/div[3]/div/label[1]/span')
elem.click()

#생년월일
elem = driver.find_element_by_id("birthday")
elem.send_keys(load_ws['B8'].value)

#빈 곳 클릭
elem = driver.find_element_by_class_name("h1")
elem.click()

#병역사항 #1
# elem = driver.find_elements_by_class_name("label")
# elem[5].click()

#병역사항 #2
elem = driver.find_element_by_link_text("군필")
elem.click()

# elem = driver.find_element_by_id("militaryRole")
# elem.send_keys(load_ws['B9'].value)

# elem = driver.find_element_by_name("military.militaryStartDate")
# elem.send_keys(load_ws['B10'].value)

# elem = driver.find_element_by_name("military.militaryEndDate")
>>>>>>> 357577a08fd34c0da72908a1c932ff307120a2f4
# elem.send_keys(load_ws['C10'].value)