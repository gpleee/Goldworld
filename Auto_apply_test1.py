<<<<<<< HEAD
from openpyxl import load_workbook
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

load_wb = load_workbook("C:/Users/GIGM/Desktop/application_info.xlsx", data_only=True)
#시트 이름으로 불러오기
load_ws = load_wb['Sheet1']

driver = webdriver.Chrome('C:/Users/GIGM/Desktop/chromedriver.exe')

driver.get(str(load_ws['B1'].value))
sleep(1)

assert "지원서" in driver.title

elem = driver.find_element_by_id("name")
elem.send_keys(load_ws['B2'].value)

elem = driver.find_element_by_id("mobile1")
elem.send_keys(load_ws['B3'].value)

elem = driver.find_element_by_id("mobile2")
elem.send_keys(load_ws['C3'].value)

elem = driver.find_element_by_id("mobile3")
elem.send_keys(load_ws['D3'].value)

elem = driver.find_element_by_id("email")
elem.send_keys(load_ws['B4'].value)

elem = driver.find_element_by_id("emailConfirm")
elem.send_keys(load_ws['B4'].value)

elem = driver.find_element_by_id("password")
elem.send_keys(load_ws['B5'].value)

elem = driver.find_element_by_id("passwordConfirm")
elem.send_keys(load_ws['B5'].value)

=======
from openpyxl import load_workbook
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

load_wb = load_workbook("C:/Users/GIGM/Desktop/application_info.xlsx", data_only=True)
#시트 이름으로 불러오기
load_ws = load_wb['Sheet1']

driver = webdriver.Chrome('C:/Users/GIGM/Desktop/chromedriver.exe')

driver.get(str(load_ws['B1'].value))
sleep(1)

assert "지원서" in driver.title

elem = driver.find_element_by_id("name")
elem.send_keys(load_ws['B2'].value)

elem = driver.find_element_by_id("mobile1")
elem.send_keys(load_ws['B3'].value)

elem = driver.find_element_by_id("mobile2")
elem.send_keys(load_ws['C3'].value)

elem = driver.find_element_by_id("mobile3")
elem.send_keys(load_ws['D3'].value)

elem = driver.find_element_by_id("email")
elem.send_keys(load_ws['B4'].value)

elem = driver.find_element_by_id("emailConfirm")
elem.send_keys(load_ws['B4'].value)

elem = driver.find_element_by_id("password")
elem.send_keys(load_ws['B5'].value)

elem = driver.find_element_by_id("passwordConfirm")
elem.send_keys(load_ws['B5'].value)

>>>>>>> 357577a08fd34c0da72908a1c932ff307120a2f4
